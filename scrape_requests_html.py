import requests
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
import random
import time
import sys
import os
from openpyxl import Workbook, load_workbook

# 獲取 HTML 頁面
def get_fetch_html(url, session, headers):
    try:
        response = session.get(url, headers=headers)
        response.raise_for_status()
        return response.text
    except Exception as e:
        error_line = sys.exc_info()[-1].tb_lineno
        print(f"錯誤發生在第 {error_line} 行：\n{str(e)}")
        return None

# 解析 list HTML 頁面，並組成 detail url
def parse_list_html(html):
    try:
        detail_urls = []
        
        soup = BeautifulSoup(html, "html.parser")
        datas = soup.find_all('a',class_="name")
        
        for data in datas:
            url = "https://ssr1.scrape.center"+data.get('href')
            detail_urls.append(url)
            
        time.sleep(random.uniform(3, 5)) 
        return detail_urls
    except Exception as e:
        error_line = sys.exc_info()[-1].tb_lineno
        print(f"錯誤發生在第 {error_line} 行：\n{str(e)}")
        return None
    
# 解析 list json 資料，並組成 detail api url
def parse_detail_html(html):
    try:
        detail_data = []
        soup = BeautifulSoup(html, "html.parser")

        # 電影名稱
        detail_data.append(soup.find("h2",class_="m-b-sm").get_text(strip=True) if soup.find("h2",class_="m-b-sm") else None)
        # 上映日期
        detail_data.append(soup.find_all("div",class_="info")[1].find('span').get_text(strip=True).replace(' 上映','') if soup.find_all("div",class_="info")[1].find('span') else None)
        # 電影評分
        detail_data.append(soup.find("p",class_="score").get_text(strip=True) if soup.find("p",class_="score") else None)
        # 電影時長
        detail_data.append(soup.find_all("div",class_="info")[0].find_all('span')[2].get_text(strip=True) if soup.find_all("div",class_="info")[0].find_all('span')[2] else None)
        # 電影簡介
        detail_data.append(soup.find("div",class_="drama").find('p').get_text(strip=True) if soup.find("div",class_="drama").find('p') else None)
        
        time.sleep(random.uniform(3, 5)) 
        
        return detail_data
    except Exception as e:
        error_line = sys.exc_info()[-1].tb_lineno
        print(f"錯誤發生在第 {error_line} 行：\n{str(e)}")
        return None

# 將資料寫入 Excel 中 
def write_to_excel(data, file_name="scrape_reques_html.xlsx"):
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["電影名稱", "上映日期", "電影評分", "電影時長", "電影簡介"])
    
    duplicate_found = False   
    for row in sheet.iter_rows(values_only=True):
        if row[0] == data[0]:
            duplicate_found = True
            break
        
    if not duplicate_found:
        sheet.append(data)
        
    workbook.save(file_name)


if __name__ == "__main__":
    session_requests = requests.session()
    ua = UserAgent()
    headers = {
        'User-Agent': ua.random
    }
    for page in range(1,99):
        list_url = f"https://ssr1.scrape.center/page/{page}"
        print(f"正在爬取電影數據網站第 {page} 頁")
        list_html = get_fetch_html(list_url, session_requests, headers)
        detail_urls = parse_list_html(list_html)
        if detail_urls != []:
            for detail_url in detail_urls:
                print(f"正在爬取電影數據網站 {detail_url}")
                detail_html = get_fetch_html(detail_url, session_requests, headers)
                detail_data = parse_detail_html(detail_html)
                print(detail_data)
                write_to_excel(detail_data)
        else:
            print(f"電影數據網站第 {page-1} 頁是最後一頁囉")
            break